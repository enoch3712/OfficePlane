"""Unit tests for the xlsx text extractor."""
import io

import pytest
from openpyxl import Workbook

from officeplane.ingestion.text_extractors.xlsx import extract_xlsx_text


def _make_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Region", "Revenue", "Growth"])
    ws.append(["NA", 1200, 0.12])
    ws.append(["EU", 900, 0.08])
    ws.append(["APAC", 750, 0.18])
    ws["B5"] = "=SUM(B2:B4)"  # formula
    ws["A5"] = "Total"

    ws2 = wb.create_sheet("Notes")
    ws2.append(["Key takeaway"])
    ws2.append(["Q3 EU softening reflects FX headwinds"])

    # Empty sheet (should be skipped)
    wb.create_sheet("Empty")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extracts_one_page_per_nonempty_sheet():
    pages = extract_xlsx_text(_make_workbook_bytes())
    titles = [p["metadata"]["sheet_name"] for p in pages]
    assert "Summary" in titles
    assert "Notes" in titles
    assert "Empty" not in titles  # filtered
    assert all(p["page_number"] >= 1 for p in pages)


def test_renders_table_as_markdown():
    pages = extract_xlsx_text(_make_workbook_bytes())
    summary = next(p for p in pages if p["metadata"]["sheet_name"] == "Summary")
    md = summary["text"]
    assert "## Summary" in md
    assert "| Region | Revenue | Growth |" in md
    assert "| NA |" in md or "| NA | 1200 |" in md or "NA" in md
    assert "EU" in md and "APAC" in md


def test_captures_formula_cells():
    pages = extract_xlsx_text(_make_workbook_bytes())
    summary = next(p for p in pages if p["metadata"]["sheet_name"] == "Summary")
    assert summary["metadata"]["formula_count"] >= 1
    assert "=SUM(B2:B4)" in summary["text"]


def test_rejects_invalid_bytes():
    with pytest.raises(ValueError):
        extract_xlsx_text(b"not an xlsx file at all")


def test_handles_completely_empty_workbook():
    wb = Workbook()
    wb.active.title = "Blank"
    buf = io.BytesIO()
    wb.save(buf)
    pages = extract_xlsx_text(buf.getvalue())
    # All sheets empty → zero pages
    assert pages == []


def test_metadata_includes_sheet_dims():
    pages = extract_xlsx_text(_make_workbook_bytes())
    summary = next(p for p in pages if p["metadata"]["sheet_name"] == "Summary")
    md = summary["metadata"]
    assert md["max_row"] >= 4
    assert md["max_column"] >= 3
    assert md["sheet_index"] == 0
