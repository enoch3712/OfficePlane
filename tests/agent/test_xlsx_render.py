"""Tests for the xlsx renderer (render_xlsx)."""
import io

from openpyxl import load_workbook

from officeplane.content_agent.renderers.workbook import parse_workbook
from officeplane.content_agent.renderers.xlsx_render import render_xlsx


def _make_wb_with_table(include_chart: bool = False):
    data = {
        "meta": {"title": "Test Workbook"},
        "sheets": [
            {
                "id": "sh1",
                "name": "Summary",
                "sections": [
                    {"type": "title", "id": "t1", "text": "BP Summary Report", "span_columns": 4},
                    {
                        "type": "table",
                        "id": "tbl1",
                        "name": "BPTable",
                        "headers": ["Month", "Systolic", "Diastolic"],
                        "rows": [
                            ["Jan", 120, 80],
                            ["Feb", 125, 82],
                            ["Mar", 118, 78],
                        ],
                        "column_formats": ["text", "number", "number"],
                        "autofilter": True,
                        "style": "TableStyleMedium2",
                    },
                ],
            }
        ],
    }
    if include_chart:
        data["sheets"][0]["sections"].append({
            "type": "chart",
            "id": "ch1",
            "chart_type": "bar",
            "title": "BP Trend",
            "data_ref": "tbl1",
            "categories_col": "Month",
            "values_col": "Systolic",
            "width_cells": 8,
            "height_cells": 15,
        })
    return parse_workbook(data)


def test_render_xlsx_produces_valid_bytes():
    wb = _make_wb_with_table()
    result = render_xlsx(wb)

    assert isinstance(result, bytes)
    assert len(result) > 1000
    # ZIP/OOXML magic bytes: PK (0x50 0x4B)
    assert result[:2] == b"PK"

    # Load with openpyxl and verify sheet is present
    loaded = load_workbook(io.BytesIO(result))
    assert "Summary" in loaded.sheetnames


def test_render_xlsx_emits_chart():
    wb = _make_wb_with_table(include_chart=True)
    result = render_xlsx(wb)

    loaded = load_workbook(io.BytesIO(result))
    ws = loaded["Summary"]
    assert len(ws._charts) >= 1, "Expected at least one chart on the worksheet"


def test_render_xlsx_handles_formulas():
    data = {
        "meta": {"title": "Formulas Test"},
        "sheets": [
            {
                "id": "sh1",
                "name": "Calcs",
                "sections": [
                    {
                        "type": "table",
                        "id": "tbl1",
                        "name": "CalcTable",
                        "headers": ["Item", "Amount"],
                        "rows": [
                            ["Alpha", 100],
                            ["Beta", 200],
                        ],
                        "column_formats": ["text", "number"],
                        "totals_row": ["Total", "=SUM(B2:B3)"],
                        "autofilter": False,
                        "style": "TableStyleMedium2",
                    }
                ],
            }
        ],
    }
    wb = parse_workbook(data)
    result = render_xlsx(wb)

    loaded = load_workbook(io.BytesIO(result), data_only=False)
    ws = loaded["Calcs"]

    # Find the totals row formula cell — it should be in the row after the data rows
    # header is row 1, data rows 2-3, totals row 4
    totals_cell = ws.cell(row=4, column=2)
    assert str(totals_cell.value) == "=SUM(B2:B3)", (
        f"Expected formula =SUM(B2:B3), got {totals_cell.value!r}"
    )


def test_render_xlsx_truncates_long_sheet_name():
    data = {
        "sheets": [
            {
                "id": "sh1",
                "name": "X" * 50,
                "sections": [],
            }
        ]
    }
    wb = parse_workbook(data)
    result = render_xlsx(wb)

    loaded = load_workbook(io.BytesIO(result))
    sheet_name = loaded.sheetnames[0]
    assert len(sheet_name) <= 31, f"Sheet name too long: {len(sheet_name)} chars"
