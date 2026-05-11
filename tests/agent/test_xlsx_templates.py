import asyncio
import json
import sys
import importlib.util
from pathlib import Path
from unittest.mock import AsyncMock, patch
import pytest


def _load(skill: str):
    p = Path(f"/app/src/officeplane/content_agent/skills/{skill}/handler.py")
    if not p.exists():
        p = Path(__file__).parents[2] / f"src/officeplane/content_agent/skills/{skill}/handler.py"
    spec = importlib.util.spec_from_file_location(f"{skill.replace('-', '_')}_handler", p)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def _seed_workbook_workspace(tmp_path: Path, ws_id: str) -> dict:
    wb = {
        "type": "workbook", "meta": {"title": "Q3"},
        "sheets": [{
            "id": "sum", "name": "Summary", "sections": [
                {"type": "title", "id": "t1", "text": "Q3 by Region"},
                {"type": "table", "id": "tab1", "name": "Revenue",
                 "headers": ["Region", "Revenue", "Growth"],
                 "rows": [["NA", 1200, 0.12], ["EU", 900, 0.08]],
                 "column_formats": ["text", "currency_usd", "percent"]},
                {"type": "table", "id": "tab2", "name": "Forecast",
                 "headers": ["Quarter", "Projected"],
                 "rows": [["Q4", 3500]],
                 "column_formats": ["text", "currency_usd"]},
                {"type": "kpi", "id": "k1", "label": "Total Q3", "value": "=SUM(B2:B3)", "format": "currency_usd"},
            ]
        }],
        "attributions": [],
    }
    workspace = tmp_path / ws_id
    workspace.mkdir(parents=True, exist_ok=True)
    (workspace / "document.json").write_text(json.dumps(wb))
    return wb


def test_save_validates(monkeypatch, tmp_path):
    mod = _load("xlsx-template-save")
    monkeypatch.setenv("CONTENT_AGENT_WORKSPACE", str(tmp_path))
    monkeypatch.setenv("OFFICEPLANE_TEMPLATES_ROOT", str(tmp_path / "templates"))
    async def _run(i): return await mod.execute(inputs=i)
    with pytest.raises(ValueError, match="workspace_id"):
        asyncio.run(_run({"name": "x"}))
    with pytest.raises(ValueError, match="name"):
        asyncio.run(_run({"workspace_id": "ws"}))


def test_save_missing_workspace(monkeypatch, tmp_path):
    mod = _load("xlsx-template-save")
    monkeypatch.setenv("CONTENT_AGENT_WORKSPACE", str(tmp_path))
    monkeypatch.setenv("OFFICEPLANE_TEMPLATES_ROOT", str(tmp_path / "templates"))
    with pytest.raises(FileNotFoundError):
        asyncio.run(mod.execute(inputs={"workspace_id": "nope", "name": "T"}))


def test_save_strips_data_rows(monkeypatch, tmp_path):
    mod = _load("xlsx-template-save")
    monkeypatch.setenv("CONTENT_AGENT_WORKSPACE", str(tmp_path))
    monkeypatch.setenv("OFFICEPLANE_TEMPLATES_ROOT", str(tmp_path / "templates"))
    _seed_workbook_workspace(tmp_path, "ws1")
    async def _run():
        with patch.object(mod, "persist_skill_invocation", new=AsyncMock(return_value=None)):
            return await mod.execute(inputs={"workspace_id": "ws1", "name": "Q3 Template"})
    r = asyncio.run(_run())
    assert r["sheet_count"] == 1
    assert r["table_count"] == 2
    saved = json.loads(Path(r["path"]).read_text())
    # Every table has empty rows
    for sh in saved["workbook"]["sheets"]:
        for sec in sh["sections"]:
            if sec["type"] == "table":
                assert sec["rows"] == []
            if sec["type"] == "kpi":
                assert sec["value"] == "=SUM(B2:B3)"  # KPI preserved


def test_apply_renders_xlsx_with_new_data(monkeypatch, tmp_path):
    save_mod = _load("xlsx-template-save")
    apply_mod = _load("xlsx-template-apply")
    monkeypatch.setenv("CONTENT_AGENT_WORKSPACE", str(tmp_path))
    monkeypatch.setenv("OFFICEPLANE_TEMPLATES_ROOT", str(tmp_path / "templates"))
    _seed_workbook_workspace(tmp_path, "ws1")

    async def _save():
        with patch.object(save_mod, "persist_skill_invocation", new=AsyncMock(return_value=None)):
            return await save_mod.execute(inputs={"workspace_id": "ws1", "name": "T"})

    r = asyncio.run(_save())
    template_id = r["template_id"]

    async def _apply():
        with patch.object(apply_mod, "persist_skill_invocation", new=AsyncMock(return_value=None)):
            return await apply_mod.execute(inputs={
                "template_id": template_id,
                "tables": {
                    "Revenue": [["NA", 1500, 0.15], ["EU", 1100, 0.10], ["APAC", 950, 0.20]],
                    "Forecast": [["Q4", 4200], ["Q1 Next", 4500]],
                },
                "title": "Q4 Run",
            })
    out = asyncio.run(_apply())
    assert out["file_path"].endswith("output.xlsx")
    assert out["title"] == "Q4 Run"
    assert out["template_id"] == template_id
    assert out["table_count"] == 2
    # Confirm real xlsx
    data = Path(out["file_path"]).read_bytes()
    assert data[:2] == b"PK"
    # Re-open and check rows landed
    from openpyxl import load_workbook
    wb = load_workbook(out["file_path"])
    ws = wb["Summary"]
    found_apac = any(
        cell.value == "APAC"
        for row in ws.iter_rows()
        for cell in row
    )
    assert found_apac


def test_apply_rejects_unknown_template(monkeypatch, tmp_path):
    apply_mod = _load("xlsx-template-apply")
    monkeypatch.setenv("CONTENT_AGENT_WORKSPACE", str(tmp_path))
    monkeypatch.setenv("OFFICEPLANE_TEMPLATES_ROOT", str(tmp_path / "templates"))
    with pytest.raises(FileNotFoundError):
        asyncio.run(apply_mod.execute(inputs={
            "template_id": "deadbeef", "tables": {"Revenue": [["NA", 1, 0.1]]}
        }))


def test_apply_rejects_unknown_table_name(monkeypatch, tmp_path):
    save_mod = _load("xlsx-template-save")
    apply_mod = _load("xlsx-template-apply")
    monkeypatch.setenv("CONTENT_AGENT_WORKSPACE", str(tmp_path))
    monkeypatch.setenv("OFFICEPLANE_TEMPLATES_ROOT", str(tmp_path / "templates"))
    _seed_workbook_workspace(tmp_path, "ws1")
    async def _save():
        with patch.object(save_mod, "persist_skill_invocation", new=AsyncMock(return_value=None)):
            return await save_mod.execute(inputs={"workspace_id": "ws1", "name": "T"})
    r = asyncio.run(_save())
    with pytest.raises(ValueError, match="no table names matched"):
        asyncio.run(apply_mod.execute(inputs={
            "template_id": r["template_id"],
            "tables": {"NonExistentTable": [["x"]]},
        }))
