"""Extract structured text from .xlsx workbooks via openpyxl.

Each worksheet becomes one "page" in the OfficePlane ingestion model.
A page's text is a markdown-flavoured rendering of the sheet's tables:
- Sheet name as H2 heading
- Each contiguous table block as a markdown table
- Cell formulas shown inline (e.g. "=SUM(B2:B10) → 1200")
- Empty rows skipped, leading/trailing whitespace trimmed

Returns the same shape as docx.py / pptx.py extractors:
list[dict] with keys: page_number, text, metadata.
"""
from __future__ import annotations

import logging
from io import BytesIO
from typing import Any

log = logging.getLogger("officeplane.ingestion.xlsx")


def extract_xlsx_text(content: bytes) -> list[dict[str, Any]]:
    """Parse .xlsx bytes into a list of page dicts (one per worksheet)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for .xlsx ingestion. "
            "Install with: pip install openpyxl"
        ) from e

    try:
        wb = load_workbook(BytesIO(content), data_only=False, read_only=False)
    except Exception as e:
        log.exception("failed to load xlsx")
        raise ValueError(f"could not parse .xlsx: {e}") from e

    pages: list[dict[str, Any]] = []
    for idx, ws in enumerate(wb.worksheets, start=1):
        markdown = _sheet_to_markdown(ws)
        if not markdown.strip():
            # Skip fully empty sheets
            continue

        # Try to load again with data_only=True to capture formula-evaluated values
        evaluated: dict[tuple[int, int], Any] = {}
        try:
            wb_eval = load_workbook(BytesIO(content), data_only=True, read_only=True)
            ws_eval = wb_eval[ws.title]
            for row_idx, row in enumerate(ws_eval.iter_rows(values_only=True), start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value is not None:
                        evaluated[(row_idx, col_idx)] = cell_value
            wb_eval.close()
        except Exception as e:
            log.warning(
                "evaluated-values load failed (formulas will show as raw): %s", e
            )

        formulas = _collect_formulas(ws, evaluated)
        named_ranges = list(wb.defined_names) if hasattr(wb, "defined_names") else []

        # Augment markdown with formula footer if any
        full_content_parts = [f"## {ws.title}", "", markdown]
        if formulas:
            full_content_parts.append("")
            full_content_parts.append("### Formulas")
            for cell_ref, fdesc in formulas[:50]:  # cap at 50 to avoid huge pages
                full_content_parts.append(f"- `{cell_ref}`: {fdesc}")

        text = "\n".join(full_content_parts)

        pages.append({
            "page_number": idx,
            "text": text,
            "metadata": {
                "sheet_name": ws.title,
                "sheet_index": idx - 1,
                "max_row": ws.max_row or 0,
                "max_column": ws.max_column or 0,
                "formula_count": len(formulas),
                "named_ranges": named_ranges[:20],  # cap
                "has_charts": bool(getattr(ws, "_charts", None)),
                "chart_count": len(getattr(ws, "_charts", []) or []),
            },
        })

    try:
        wb.close()
    except Exception:
        pass

    return pages


def _sheet_to_markdown(ws: Any) -> str:
    """Render a worksheet to markdown.

    Identifies header rows and renders contiguous blocks as tables.
    Simple heuristic: a contiguous run of non-empty rows is a "block";
    the first row of each block is treated as headers.
    """
    if ws.max_row is None or ws.max_row < 1:
        return ""

    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rendered = [_render_cell(v) for v in row]
        # Strip trailing empties
        while rendered and rendered[-1] == "":
            rendered.pop()
        rows.append(rendered)

    # Split into blocks by blank rows
    blocks: list[list[list[str]]] = []
    current: list[list[str]] = []
    for r in rows:
        if any(c.strip() for c in r):
            current.append(r)
        else:
            if current:
                blocks.append(current)
                current = []
    if current:
        blocks.append(current)

    md_parts: list[str] = []
    for block in blocks:
        if not block:
            continue
        # Single-cell blocks (probably titles) — emit as plain text
        if len(block) == 1 and len(block[0]) <= 1:
            md_parts.append(block[0][0] if block[0] else "")
            continue
        max_cols = max(len(r) for r in block)
        # Normalise width
        norm = [r + [""] * (max_cols - len(r)) for r in block]
        # Treat first row as header
        header = norm[0]
        body = norm[1:]
        # Build markdown table
        md_parts.append("| " + " | ".join(_md_escape(c) for c in header) + " |")
        md_parts.append("| " + " | ".join(["---"] * max_cols) + " |")
        for row in body:
            md_parts.append("| " + " | ".join(_md_escape(c) for c in row) + " |")
        md_parts.append("")  # blank line between blocks

    return "\n".join(md_parts).strip()


def _render_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        # Avoid scientific notation for common ranges
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        return f"{v:g}"
    return str(v)


def _md_escape(s: str) -> str:
    return s.replace("|", "\\|").replace("\n", " ").strip()


def _collect_formulas(
    ws: Any, evaluated: dict[tuple[int, int], Any]
) -> list[tuple[str, str]]:
    """Return list of (cell_ref, "formula → evaluated_value") for formula cells."""
    out: list[tuple[str, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                ev = evaluated.get((cell.row, cell.column))
                desc = f"`{v}`"
                if ev is not None and not (isinstance(ev, str) and ev.startswith("=")):
                    desc += f" → {_render_cell(ev)}"
                out.append((cell.coordinate, desc))
    return out
