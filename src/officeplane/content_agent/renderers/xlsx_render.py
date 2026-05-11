"""Render Workbook → .xlsx bytes via openpyxl."""
from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from officeplane.content_agent.renderers.workbook import (
    BlankSection, ChartSection, KpiSection, Sheet, SubtitleSection, TableSection,
    TextSection, TitleSection, Workbook,
)

log = logging.getLogger("officeplane.renderers.xlsx")

FORMAT_CODES = {
    "text": "@", "number": "0.00", "integer": "0",
    "currency_usd": '"$"#,##0.00', "currency_eur": '"€"#,##0.00',
    "percent": "0.0%", "date": "yyyy-mm-dd", "datetime": "yyyy-mm-dd hh:mm",
}
PHOSPHOR = "5EFCAB"


def render_xlsx(wb: Workbook, *, workspace_dir: Path | None = None) -> bytes:
    obx = OpenpyxlWorkbook()
    obx.remove(obx.active)
    for sheet in (wb.sheets or [Sheet(id="default", name="Sheet1")]):
        _render_sheet(obx, sheet)
    if not obx.sheetnames:
        obx.create_sheet("Sheet1")
    obx.properties.title = wb.meta.title
    if wb.meta.author:
        obx.properties.creator = wb.meta.author
    if wb.meta.description:
        obx.properties.description = wb.meta.description
    buf = io.BytesIO()
    obx.save(buf)
    return buf.getvalue()


def _render_sheet(obx: OpenpyxlWorkbook, sheet: Sheet) -> None:
    ws = obx.create_sheet(sheet.name[:31])
    row = 1
    widths: dict[int, int] = {}
    rendered_tables: dict[str, dict[str, Any]] = {}

    for section in sheet.sections:
        if isinstance(section, TitleSection):
            row = _render_title(ws, section, row, widths)
        elif isinstance(section, SubtitleSection):
            row = _render_subtitle(ws, section, row, widths)
        elif isinstance(section, TextSection):
            row = _render_text(ws, section, row, widths)
        elif isinstance(section, BlankSection):
            row += 1
        elif isinstance(section, TableSection):
            info = _render_table(ws, section, row, widths)
            rendered_tables[section.id] = info
            row = info["next_row"]
        elif isinstance(section, ChartSection):
            row = _render_chart(ws, section, row, rendered_tables)
        elif isinstance(section, KpiSection):
            row = _render_kpi(ws, section, row, widths)

    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(8, w + 2))
    for letter, w in sheet.column_widths.items():
        ws.column_dimensions[letter].width = w

    if sheet.freeze_header:
        first_header = _find_first_table_header_row(sheet)
        if first_header:
            ws.freeze_panes = f"A{first_header + 1}"


def _measure(widths: dict[int, int], col: int, value: Any) -> None:
    s = str(value) if value is not None else ""
    widths[col] = max(widths.get(col, 0), len(s))


def _render_title(ws, section: TitleSection, row: int, widths: dict[int, int]) -> int:
    cell = ws.cell(row=row, column=1, value=section.text)
    cell.font = Font(size=18, bold=True, color=PHOSPHOR)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if section.span_columns > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=section.span_columns)
    ws.row_dimensions[row].height = 28
    _measure(widths, 1, section.text)
    return row + 2


def _render_subtitle(ws, section: SubtitleSection, row: int, widths: dict[int, int]) -> int:
    cell = ws.cell(row=row, column=1, value=section.text)
    cell.font = Font(size=13, bold=True, color="9CA3AF")
    _measure(widths, 1, section.text)
    return row + 1


def _render_text(ws, section: TextSection, row: int, widths: dict[int, int]) -> int:
    cell = ws.cell(row=row, column=1, value=section.text)
    cell.font = Font(size=11)
    _measure(widths, 1, section.text)
    return row + 1


def _coerce(v: Any) -> Any:
    if isinstance(v, str) and v.startswith("="):
        return v
    if isinstance(v, (int, float, bool)) or v is None:
        return v
    return str(v)


def _render_table(ws, section: TableSection, row: int, widths: dict[int, int]) -> dict[str, Any]:
    n_cols = max(len(section.headers), 1)
    for col_idx, h in enumerate(section.headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        c.alignment = Alignment(horizontal="left", vertical="center")
        _measure(widths, col_idx, h)
    header_row = row
    row += 1
    formats = list(section.column_formats) + ["text"] * max(0, n_cols - len(section.column_formats))
    for r in section.rows:
        for col_idx, value in enumerate(r[:n_cols], start=1):
            c = ws.cell(row=row, column=col_idx, value=_coerce(value))
            fmt = formats[col_idx - 1] if col_idx - 1 < len(formats) else "text"
            c.number_format = FORMAT_CODES.get(fmt, "General")
            _measure(widths, col_idx, value)
        row += 1
    last_data_row = row - 1
    if section.totals_row:
        for col_idx, value in enumerate(section.totals_row[:n_cols], start=1):
            c = ws.cell(row=row, column=col_idx, value=_coerce(value))
            c.font = Font(bold=True)
            c.border = Border(top=Side(style="thin", color="666666"))
            fmt = formats[col_idx - 1] if col_idx - 1 < len(formats) else "text"
            c.number_format = FORMAT_CODES.get(fmt, "General")
            _measure(widths, col_idx, value)
        row += 1
    end_row = row - 1
    if section.autofilter and n_cols > 0 and end_row > header_row:
        try:
            ref = f"A{header_row}:{get_column_letter(n_cols)}{end_row}"
            safe = "".join(c if c.isalnum() else "_" for c in section.name)[:31] or f"Table_{section.id}"
            tbl = Table(displayName=safe, ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name=section.style, showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False)
            ws.add_table(tbl)
        except Exception as e:
            log.warning("table %s failed: %s", section.id, e)
    return {"header_row": header_row, "data_start_row": header_row + 1,
            "data_end_row": last_data_row, "n_cols": n_cols,
            "headers": list(section.headers), "next_row": row + 1}


def _render_chart(ws, section: ChartSection, row: int, tables: dict[str, dict[str, Any]]) -> int:
    info = tables.get(section.data_ref)
    if not info:
        log.warning("chart %s missing data_ref %s", section.id, section.data_ref)
        return row + 1
    cats = _resolve_col(section.categories_col, info["headers"])
    vals = _resolve_col(section.values_col, info["headers"])
    if cats is None or vals is None:
        return row + 1
    cls = {"bar": BarChart, "column": BarChart, "line": LineChart,
           "pie": PieChart, "scatter": ScatterChart}.get(section.chart_type, BarChart)
    chart = cls()
    chart.title = section.title or section.id
    if hasattr(chart, "type") and section.chart_type == "column":
        chart.type = "col"
    chart.style = 11
    values = Reference(ws, min_col=vals, min_row=info["header_row"],
                       max_col=vals, max_row=info["data_end_row"])
    categories = Reference(ws, min_col=cats, min_row=info["data_start_row"],
                           max_col=cats, max_row=info["data_end_row"])
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart.width = section.width_cells * 1.7
    chart.height = section.height_cells * 0.6
    anchor = f"{get_column_letter(info['n_cols'] + 2)}{row}"
    ws.add_chart(chart, anchor)
    return row + max(15, section.height_cells)


def _render_kpi(ws, section: KpiSection, row: int, widths: dict[int, int]) -> int:
    label = ws.cell(row=row, column=1, value=section.label)
    label.font = Font(size=11, bold=True, color="9CA3AF")
    value = ws.cell(row=row, column=2, value=_coerce(section.value))
    value.font = Font(size=18, bold=True, color=PHOSPHOR)
    value.alignment = Alignment(horizontal="left", vertical="center")
    value.number_format = FORMAT_CODES.get(section.format, "General")
    value.border = Border(left=Side(style="medium", color=PHOSPHOR),
                          bottom=Side(style="thin", color="333333"))
    ws.row_dimensions[row].height = 26
    _measure(widths, 1, section.label)
    _measure(widths, 2, section.value)
    return row + 2


def _resolve_col(spec: str, headers: list[str]) -> int | None:
    if not spec:
        return None
    if spec.isdigit():
        idx = int(spec)
        return idx + 1 if 0 <= idx < len(headers) else None
    for i, h in enumerate(headers, start=1):
        if h.strip().lower() == spec.strip().lower():
            return i
    return None


def _find_first_table_header_row(sheet: Sheet) -> int | None:
    row = 1
    for sec in sheet.sections:
        if isinstance(sec, TitleSection):
            row += 2
        elif isinstance(sec, (SubtitleSection, TextSection)):
            row += 1
        elif isinstance(sec, BlankSection):
            row += 1
        elif isinstance(sec, TableSection):
            return row
        elif isinstance(sec, (ChartSection, KpiSection)):
            row += 2
    return None
